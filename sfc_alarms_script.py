import pprint
import datetime
import re
import sys
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment


def read_me():
    print()
    print('***************************************        SFC ALARMS ANALYSIS         ***************************************')
    print('1. This script creates a summary report for all alarms observed in a given period.')
    print('2. The provided .xlsx file MUST have SuR,4G,3G and 2G alarms worksheets for this script to work')
    print(
        '3. It creates a new "_updated_[date].xlsx" file on the directory the script is run')
    print('4. A new worksheet "SuR Analysis" is created in the above document where further analysis can be conducted')
    print('5. Inputs: minimum unavailability threshold, alarms dates')
    print('6. Sites with unavailability mins above the minimum threshold will have their alarms summarized for the given dates')
    print('                                              Developer:   Peter Kimani \n                                              Contact:     peteratnokia@gmail.com\n                                              Version: 1.0')
    print('******************************************************************************************************************')
    print()


def date_input_and_Validations():
    '''
    Date input and validations
    Return a list of dates for SuR Consideration
    '''
    dates = []
    while True:
        try:
            minimum_threshold = float(
                input('INPUT:\tEnter the minimum threshold of unavailability in mins : '))
            break
        except Exception as e:
            print('WARNING:\t{}'.format(e))
    while True:
        choice = str(input(
            'INPUT:\tDo you want to perform SuR Alarm Analysis for the last [0,1...,n] dates? y/n : '))
        if choice.lower() == 'y' or choice.lower() == 'n':
            if choice.lower() == 'y':
                while True:
                    try:
                        no_of_days = int(
                            input('INPUT:\tAnalyze alarms for the last ? days: '))
                        if no_of_days == 0:
                            print(
                                'INFO:\t Will analyze alarms for today\'s date by default')
                            no_of_days = 1  # making default value of 1
                        break
                    except:
                        print('WARNING:\tEnter an integer value !!!')
                for day in range(0, no_of_days):
                    date = str(datetime.date.today() -
                               datetime.timedelta(days=day))
                    year = date[0:4]
                    month = date[5:7]
                    day = date[8:10]
                    date_Format = '{}-{}-{}'.format(year, month, day)
                    dates.append(date_Format)
                break
            else:
                while True:
                    print('\nINFO:\tChoose a specific date for alarm analysis\n')
                    year = str(input('INPUT:\tEnter the year          ::: '))
                    if year.isdecimal():
                        if int(year) > 2020 and int(year) < 2025:
                            while True:
                                search_date = str(
                                    input('INPUT:\tEnter the date (DD-MM) ::: '))
                                if (search_date[0:2].isdecimal()) and len(search_date[0:2]) == 2 and (int(search_date[0:2]) > 0 and int(search_date[0:2]) < 32):
                                    if search_date[2:3] == '-':
                                        if (search_date[3:5].isdecimal()) and len(search_date[3:5]) == 2 and (int(search_date[3:5]) > 0 and int(search_date[3:5]) < 13):
                                            dates = [
                                                '{}-{}-{}'.format(year, search_date[3:5], search_date[0:2])]
                                            break
                                        else:
                                            print(
                                                'WARNING:\tEnter the right format of the month portion of the date.\n\tIt should be 2 digits e.g 01 and be between 1-12 (DD-MM)')
                                    else:
                                        print(
                                            'WARNING:\tDate separator must be a dash "-"(DD-MM)')
                                else:
                                    print(
                                        'WARNING: Enter the right format of the day portion of the date.\n\tIt should be 2 digits e.g 01 and be between 1-31 (DD-MM)')
                            break
                        else:
                            print('WARNING: Base year: 2021-2025')
                    else:
                        print('WARNING: Enter a valid year e.g. 2021')
                break
        else:
            print('WARNING: Please enter the correct choice (y/n)')

    return dates, minimum_threshold


def alarms_Analysis(minimum_threshold, week_dates, df_sur, df_4G, df_3G, df_2G):
    '''
    Returns a Dict data structure for alarms analysis
    '''
    alarms_analysis_dic = {}
    print('INFO:\tAnalyzing Alarms ...')
    for ind in df_sur.index:
        try:
            if df_sur['Unavailable_Mins'][ind] > minimum_threshold:
                try:
                    site_id = str(int(df_sur['Physical_id'][ind]))
                except:
                    site_id = 'NULL'

                if df_sur['Technology'][ind] == '4G':
                    alarms_analysis_dic.setdefault(
                        df_sur['Region'][ind], {})  # --- Region
                    alarms_analysis_dic[df_sur['Region'][ind]].setdefault(
                        df_sur['Technology'][ind], {})  # --- Tech i.e 2G,3G,4G
                    alarms_analysis_dic[df_sur['Region'][ind]][df_sur['Technology'][ind]].setdefault(
                        df_sur['Node_Name'][ind], {})  # ---    Node-Name-BSC/RNC
                    alarms_analysis_dic[df_sur['Region'][ind]][df_sur['Technology'][ind]
                                                               ][df_sur['Node_Name'][ind]].setdefault(site_id, {})  # --- Site-id

                    alarms_analysis_dic[df_sur['Region'][ind]][df_sur['Technology'][ind]][df_sur['Node_Name'][ind]][site_id].setdefault(
                        df_sur['Cell_name'][ind], {df_sur['Unavailable_Mins'][ind]: {}})  # --- Cell-id & Unavailability

                    alarms = {}

                    for alarm_index in df_4G.index:
                        if re.search(site_id, str(df_4G['CONTROLLING_OBJECT'][alarm_index])) and df_4G['ALARM_TIME'][alarm_index].strftime('%Y-%m-%d') in week_dates:
                            alarms.setdefault(
                                df_4G['ALARM_TEXT'][alarm_index], {})
                            alarms[df_4G['ALARM_TEXT'][alarm_index]].setdefault(
                                df_4G['SUPPLEMENTARY_INFO'][alarm_index], 0)
                            alarms[df_4G['ALARM_TEXT'][alarm_index]][df_4G['SUPPLEMENTARY_INFO'][alarm_index]
                                                                     ] = alarms[df_4G['ALARM_TEXT'][alarm_index]].setdefault(df_4G['SUPPLEMENTARY_INFO'][alarm_index], 0)+1

                            alarms_analysis_dic[df_sur['Region'][ind]][df_sur['Technology'][ind]][df_sur['Node_Name'][ind]
                                                                                                  ][site_id][df_sur['Cell_name'][ind]][df_sur['Unavailable_Mins'][ind]] = alarms  # --- Alarms

                elif df_sur['Technology'][ind] == '3G':
                    alarms_analysis_dic.setdefault(
                        df_sur['Region'][ind], {})  # --- Region
                    alarms_analysis_dic[df_sur['Region'][ind]].setdefault(
                        df_sur['Technology'][ind], {})  # --- Tech i.e 2G,3G,4G
                    alarms_analysis_dic[df_sur['Region'][ind]][df_sur['Technology'][ind]].setdefault(
                        df_sur['Node_Name'][ind], {})  # ---    Node-Name-BSC/RNC
                    alarms_analysis_dic[df_sur['Region'][ind]][df_sur['Technology'][ind]
                                                               ][df_sur['Node_Name'][ind]].setdefault(site_id, {})  # --- Site-id

                    alarms_analysis_dic[df_sur['Region'][ind]][df_sur['Technology'][ind]][df_sur['Node_Name'][ind]][site_id].setdefault(
                        df_sur['Cell_name'][ind], {df_sur['Unavailable_Mins'][ind]: {}})  # --- Cell-id & Unavailability

                    alarms = {}

                    for alarm_index in df_3G.index:
                        if re.search(site_id, str(df_3G['NAME'][alarm_index])) and df_3G['ALARM_TIME'][alarm_index].strftime('%Y-%m-%d') in week_dates:
                            alarms.setdefault(
                                df_3G['ALARM_TEXT'][alarm_index], {})
                            alarms[df_3G['ALARM_TEXT'][alarm_index]].setdefault(
                                df_3G['SUPPLEMENTARY_INFO'][alarm_index], 0)
                            alarms[df_3G['ALARM_TEXT'][alarm_index]][df_3G['SUPPLEMENTARY_INFO'][alarm_index]
                                                                     ] = alarms[df_3G['ALARM_TEXT'][alarm_index]].setdefault(df_3G['SUPPLEMENTARY_INFO'][alarm_index], 0)+1

                            alarms_analysis_dic[df_sur['Region'][ind]][df_sur['Technology'][ind]][df_sur['Node_Name'][ind]
                                                                                                  ][site_id][df_sur['Cell_name'][ind]][df_sur['Unavailable_Mins'][ind]] = alarms  # --- Alarms

                elif df_sur['Technology'][ind] == '2G':
                    alarms_analysis_dic.setdefault(
                        df_sur['Region'][ind], {})  # --- Region
                    alarms_analysis_dic[df_sur['Region'][ind]].setdefault(
                        df_sur['Technology'][ind], {})  # --- Tech i.e 2G,3G,4G
                    alarms_analysis_dic[df_sur['Region'][ind]][df_sur['Technology'][ind]].setdefault(
                        df_sur['Node_Name'][ind], {})  # ---    Node-Name-BSC/RNC
                    alarms_analysis_dic[df_sur['Region'][ind]][df_sur['Technology'][ind]
                                                               ][df_sur['Node_Name'][ind]].setdefault(site_id, {})  # --- Site-id

                    alarms_analysis_dic[df_sur['Region'][ind]][df_sur['Technology'][ind]][df_sur['Node_Name'][ind]][site_id].setdefault(
                        df_sur['Cell_name'][ind], {df_sur['Unavailable_Mins'][ind]: {}})  # --- Cell-id & Unavailability

                    alarms = {}

                    for alarm_index in df_2G.index:
                        # Looking for a match in specific tech Alarms Worksheet on the corresponding dates for a site id
                        if re.search(site_id, str(df_2G['NAME'][alarm_index])) and df_2G['ALARM_TIME'][alarm_index].strftime('%Y-%m-%d') in week_dates:
                            alarms.setdefault(
                                df_2G['ALARM_TEXT'][alarm_index], 0)
                            alarms[df_2G['ALARM_TEXT'][alarm_index]] = alarms.setdefault(
                                df_2G['ALARM_TEXT'][alarm_index], 0)+1
                            #--- Eliminating Supplimentary Info from the data as it does not provide any iseful info------#
                            # alarms[df_2G['ALARM_TEXT'][alarm_index]].setdefault(df_2G['SUPPLEMENTARY_INFO'][alarm_index],0)
                            # alarms[df_2G['ALARM_TEXT'][alarm_index]][df_2G['SUPPLEMENTARY_INFO'][alarm_index]]= alarms[df_2G['ALARM_TEXT'][alarm_index]].setdefault(df_2G['SUPPLEMENTARY_INFO'][alarm_index],0)+1

                            alarms_analysis_dic[df_sur['Region'][ind]][df_sur['Technology'][ind]][df_sur['Node_Name'][ind]
                                                                                                  ][site_id][df_sur['Cell_name'][ind]][df_sur['Unavailable_Mins'][ind]] = alarms  # --- Alarms
        except Exception as e:
            print('WARNING:\t{} {}'.format(e, df_sur['Unavailable_Mins'][ind]))
    print('INFO:\tAnalyzing Alarms Complete')
    return alarms_analysis_dic


def main(excel_filename):
    '''
    Main
    '''
    read_me()
    try:
        print('INFO:\tLoading "{}" ...'.format(excel_filename))
        wb = load_workbook(excel_filename)
        for sheetname in wb.sheetnames:
            worksheet = wb[sheetname]
            if re.search('SuR', sheetname, re.IGNORECASE):
                # Only creating a dataframe once when df variable does not exist. Use case several sheetnames with keyword "SuR"
                try:
                    if df_sur:
                        pass
                except:
                    print('INFO:\tReading SuR into a DataFrame ...')
                    df_sur = pd.read_excel(
                        excel_filename, sheet_name=sheetname, na_filter=False)
            elif re.search('4G', sheetname, re.IGNORECASE):
                # Only creating a dataframe once when df variable does not exist. Use case several sheetnames with keyword "4G"
                try:
                    if df_4G:
                        pass
                except:
                    print('INFO:\tReading 4G Alarms into a DataFrame ...')
                    df_4G = pd.read_excel(
                        excel_filename, sheet_name=sheetname, na_filter=False)
            elif re.search('3G', sheetname, re.IGNORECASE):
                # Only creating a dataframe once when df variable does not exist. Use case several sheetnames with keyword "3G"
                try:
                    if df_3G:
                        pass
                except:
                    print('INFO:\tReading 3G Alarms into a DataFrame ...')
                    df_3G = pd.read_excel(
                        excel_filename, sheet_name=sheetname, na_filter=False)
            elif re.search('2G', sheetname, re.IGNORECASE):
                # Only creating a dataframe once when df variable does not exist. Use case several sheetnames with keyword "2G"
                try:
                    if df_2G:
                        pass
                except:
                    print('INFO:\tReading 2G Alarms into a DataFrame ...')
                    df_2G = pd.read_excel(
                        excel_filename, sheet_name=sheetname, na_filter=False)
        week_dates, minimum_threshold = date_input_and_Validations()
        alarms_analysis_dic = alarms_Analysis(
            minimum_threshold, week_dates, df_sur, df_4G, df_3G, df_3G)
        try:
            print('INFO:\tTrying to open "SuR Analysis" Worksheet ...')
            ws = wb['SuR Analysis']
        except Exception as e:
            print('INFO:\t{}'.format(e))
            print('INFO:\tCreating "SuR Analysis" Worksheet ...')
            wb.create_sheet('SuR Analysis')
            print('INFO:\tOpening "SuR Analysis" Worksheet ...')
            ws = wb['SuR Analysis']

        filters = ['Technology', 'Region', 'Node_Name', 'Physical_id', 'Cell_name',
                   'Unavailable_Mins', 'Alarms Analysis', 'Root_Cause', 'Expert']
        filters_cells = ['A1', 'B1', 'C1', 'D1', 'E1', 'F1', 'G1', 'H1', 'I1']
        print('INFO:\tWriting and formating headers ...')
        #------Writing and formating the headers/filters----------------#
        for index in range(0, len(filters), 1):
            ws[filters_cells[index]].value = filters[index]
            ws[filters_cells[index]].fill = PatternFill(
                "solid", fgColor="ffff00")  # Making FG Color Yellow
        ws.column_dimensions['A'].width = 10  # --- Tech
        ws.column_dimensions['B'].width = 13  # --- Region
        ws.column_dimensions['C'].width = 11  # --- Node_Name
        ws.column_dimensions['D'].width = 10  # --- Physical_id
        ws.column_dimensions['E'].width = 46  # --- Cell_name
        ws.column_dimensions['F'].width = 14  # --- Unavailable_Mins
        ws.column_dimensions['G'].width = 70  # --- Alarms
        ws.column_dimensions['H'].width = 14  # --- Root_Cause
        ws.column_dimensions['I'].width = 10  # --- Expert
        #------Writing the Summarized Dict Data Structure data to the Worksheet----------#
        text_Font = Font(name='Calibri', sz=8.0)
        row = 2
        for region in alarms_analysis_dic.keys():
            for tech in alarms_analysis_dic[region].keys():
                for node in alarms_analysis_dic[region][tech].keys():
                    for site in alarms_analysis_dic[region][tech][node].keys():
                        for cells in alarms_analysis_dic[region][tech][node][site].keys():
                            for unavailability, alarms in alarms_analysis_dic[region][tech][node][site][cells].items():
                                print(
                                    'INFO:\tWriting and formating data for "{}"'.format(cells))

                                ws['A'+str(row)].value = tech
                                ws['A'+str(row)].font = text_Font
                                ws['A'+str(row)].alignment = Alignment(horizontal="left",
                                                                       vertical="center", wrap_text=False)

                                ws['B'+str(row)].value = region
                                ws['B'+str(row)].font = text_Font
                                ws['B'+str(row)].alignment = Alignment(horizontal="left",
                                                                       vertical="center", wrap_text=False)

                                ws['C'+str(row)].value = node
                                ws['C'+str(row)].font = text_Font
                                ws['C'+str(row)].alignment = Alignment(horizontal="left",
                                                                       vertical="center", wrap_text=False)

                                try:
                                    ws['D'+str(row)].value = int(site)
                                except:
                                    ws['D'+str(row)].value = site
                                ws['D'+str(row)].font = text_Font
                                ws['D'+str(row)].alignment = Alignment(horizontal="left",
                                                                       vertical="center", wrap_text=False)

                                ws['E'+str(row)].value = cells
                                ws['E'+str(row)].font = text_Font
                                ws['E'+str(row)].alignment = Alignment(horizontal="left",
                                                                       vertical="center", wrap_text=False)

                                ws['F'+str(row)].value = unavailability
                                ws['F'+str(row)].font = text_Font
                                ws['F'+str(row)].alignment = Alignment(horizontal="left",
                                                                       vertical="center", wrap_text=False)

                                ws['G'+str(row)].value = pprint.pformat(alarms)
                                ws['G'+str(row)].font = text_Font
                                ws['G'+str(row)].alignment = Alignment(horizontal="left",
                                                                       vertical="center", wrap_text=True)

                                row += 1
        #------Making column headers filters----------------#
        ws.auto_filter.ref = ws.dimensions
        #------Freezing the top row to allow scrolling------------#
        ws.freeze_panes = "B2"

        # ---Save the workbook----
        updated_file_name = excel_filename[0:(len(excel_filename)-5)]+'_updated_on_{}'.format(
            week_dates[0])+excel_filename[(len(excel_filename)-5):]  # Appends 'updated' to the initial filename
        print('INFO:\tSaving "{}" ...'.format(updated_file_name))
        wb.save(updated_file_name)
    except Exception as e:
        print('ERROR:\t{}\n\tClosing the read\write process'.format(e))
        updated_file_name = excel_filename
    finally:
        print('INFO:\tClosing "{}"'.format(updated_file_name))
        try:
            wb.close()  # Closing the filename if it was already open
        except:
            pass  # Neglecting any exceptions if the file was not already opened


if __name__ == '__main__':

    if len(sys.argv) == 1:
        print('ERROR:\tAdd the filename argument i.e "script.exe path_to_filename.xlsx"')
    else:
        excel_filename = ''
        for arg in range(1, len(sys.argv), 1):
            excel_filename = excel_filename + sys.argv[arg] + ' '
        # Removing the extra space at the end of the filename
        excel_filename = excel_filename[:(len(excel_filename)-1)]
        if excel_filename[(len(excel_filename)-5):] == '.xlsx':
            main(excel_filename)
        else:
            print('ERROR:\tAdd the ".xlsx" file extension')
